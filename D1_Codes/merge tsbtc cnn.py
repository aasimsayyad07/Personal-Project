import csv
import pandas
import xlsxwriter
import openpyxl

workbook1=openpyxl.load_workbook(filename='cnn_data_and_category.xlsx')
wsheet1=workbook1['Sheet1']
print('start')
col_gap=3
for m in range(2,11):
    print("start "+str(m))
    print('')
    col_gap=col_gap+3
    csv_file = pandas.read_csv('task2_'+str(m)+'ary.csv',index_col="img_id")
    wsheet1.insert_cols(1,col_gap)

    for i in range(2,10018):
        img_name=wsheet1.cell(row=i,column=1002+col_gap).value
        
        for j in range(1,col_gap+1):
            wsheet1.cell(row=i,column=j,value=csv_file['f'+str(j)][img_name])
        print(i)

    
    
    workbook1.save(filename='tsbtc'+str(m)+'ary_and_cnn.xlsx')
    
    print("done"+str(m))
    wsheet1.delete_cols(1,col_gap)
